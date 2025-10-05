"""
Advanced financial reporting services.
"""
from decimal import Decimal
from datetime import datetime, date, timedelta
from typing import Dict, List, Any, Optional
from django.db.models import Q, Sum, Count, Avg, F, Case, When, Value
from django.db.models.functions import Coalesce, Extract, TruncMonth, TruncQuarter, TruncYear
from django.utils.translation import gettext_lazy as _

from modules.accounting.models import Account, JournalEntry, JournalEntryLine
from modules.invoicing.models import Invoice, Payment
from modules.sales.models import SalesOrder
from modules.purchasing.models import PurchaseOrder
from core.companies.models import Company


class AdvancedFinancialReportService:
    """Service for generating advanced financial reports."""
    
    def __init__(self, company: Company):
        self.company = company
    
    def generate_detailed_pnl(
        self,
        start_date: date,
        end_date: date,
        comparison_period: Optional[tuple] = None,
        include_budget: bool = False
    ) -> Dict[str, Any]:
        """Generate detailed P&L with comparisons and variance analysis."""
        
        # Get current period data
        current_data = self._get_pnl_data(start_date, end_date)
        
        result = {
            'period': {
                'start_date': start_date,
                'end_date': end_date,
                'data': current_data
            },
            'comparison': None,
            'variance': None,
            'budget': None
        }
        
        # Add comparison period if provided
        if comparison_period:
            comp_start, comp_end = comparison_period
            comparison_data = self._get_pnl_data(comp_start, comp_end)
            result['comparison'] = {
                'start_date': comp_start,
                'end_date': comp_end,
                'data': comparison_data
            }
            
            # Calculate variance
            result['variance'] = self._calculate_variance(current_data, comparison_data)
        
        # Add budget comparison if requested
        if include_budget:
            # TODO: Implement budget comparison when budget module is available
            pass
        
        return result
    
    def _get_pnl_data(self, start_date: date, end_date: date) -> Dict[str, Any]:
        """Get P&L data for a specific period."""
        
        # Get revenue accounts (7xxx)
        revenue_accounts = Account.objects.filter(
            company=self.company,
            code__startswith='7',
            is_active=True
        )
        
        # Get expense accounts (6xxx)
        expense_accounts = Account.objects.filter(
            company=self.company,
            code__startswith='6',
            is_active=True
        )
        
        # Calculate revenue
        revenue_data = self._get_account_balances(
            revenue_accounts,
            start_date,
            end_date,
            'CREDIT'
        )
        
        # Calculate expenses
        expense_data = self._get_account_balances(
            expense_accounts,
            start_date,
            end_date,
            'DEBIT'
        )
        
        # Calculate totals
        total_revenue = sum(item['balance'] for item in revenue_data.values())
        total_expenses = sum(item['balance'] for item in expense_data.values())
        net_income = total_revenue - total_expenses
        
        return {
            'revenue': {
                'accounts': revenue_data,
                'total': total_revenue
            },
            'expenses': {
                'accounts': expense_data,
                'total': total_expenses
            },
            'net_income': net_income,
            'gross_margin': (total_revenue - total_expenses) / total_revenue * 100 if total_revenue > 0 else 0
        }
    
    def _get_account_balances(
        self,
        accounts,
        start_date: date,
        end_date: date,
        balance_type: str
    ) -> Dict[str, Dict[str, Any]]:
        """Get account balances for a period."""
        
        result = {}
        
        for account in accounts:
            # Get journal entry lines for this account in the period
            lines = JournalEntryLine.objects.filter(
                account=account,
                journal_entry__date__range=[start_date, end_date],
                journal_entry__state='POSTED'
            )
            
            if balance_type == 'DEBIT':
                balance = lines.aggregate(
                    total=Coalesce(Sum('debit_amount'), Decimal('0'))
                )['total']
            else:  # CREDIT
                balance = lines.aggregate(
                    total=Coalesce(Sum('credit_amount'), Decimal('0'))
                )['total']
            
            if balance > 0:
                result[account.code] = {
                    'name': account.name,
                    'balance': balance,
                    'account_type': account.account_type.name
                }
        
        return result
    
    def _calculate_variance(
        self,
        current_data: Dict[str, Any],
        comparison_data: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Calculate variance between two periods."""
        
        variance = {}
        
        # Revenue variance
        current_revenue = current_data['revenue']['total']
        comparison_revenue = comparison_data['revenue']['total']
        revenue_variance = current_revenue - comparison_revenue
        revenue_variance_pct = (revenue_variance / comparison_revenue * 100) if comparison_revenue > 0 else 0
        
        variance['revenue'] = {
            'absolute': revenue_variance,
            'percentage': revenue_variance_pct
        }
        
        # Expense variance
        current_expenses = current_data['expenses']['total']
        comparison_expenses = comparison_data['expenses']['total']
        expense_variance = current_expenses - comparison_expenses
        expense_variance_pct = (expense_variance / comparison_expenses * 100) if comparison_expenses > 0 else 0
        
        variance['expenses'] = {
            'absolute': expense_variance,
            'percentage': expense_variance_pct
        }
        
        # Net income variance
        current_net = current_data['net_income']
        comparison_net = comparison_data['net_income']
        net_variance = current_net - comparison_net
        net_variance_pct = (net_variance / comparison_net * 100) if comparison_net != 0 else 0
        
        variance['net_income'] = {
            'absolute': net_variance,
            'percentage': net_variance_pct
        }
        
        return variance
    
    def generate_cash_flow_analysis(
        self,
        start_date: date,
        end_date: date,
        forecast_days: int = 30
    ) -> Dict[str, Any]:
        """Generate detailed cash flow analysis with forecasting."""
        
        # Get cash accounts
        cash_accounts = Account.objects.filter(
            company=self.company,
            code__in=['5111', '5112', '5113'],  # Cash, Bank, Short-term investments
            is_active=True
        )
        
        # Operating cash flow
        operating_cf = self._calculate_operating_cash_flow(start_date, end_date)
        
        # Investing cash flow
        investing_cf = self._calculate_investing_cash_flow(start_date, end_date)
        
        # Financing cash flow
        financing_cf = self._calculate_financing_cash_flow(start_date, end_date)
        
        # Net cash flow
        net_cf = operating_cf + investing_cf + financing_cf
        
        # Cash position
        cash_position = self._get_cash_position(cash_accounts, end_date)
        
        # Forecast
        forecast = self._forecast_cash_flow(forecast_days) if forecast_days > 0 else None
        
        return {
            'period': {
                'start_date': start_date,
                'end_date': end_date
            },
            'operating_cash_flow': operating_cf,
            'investing_cash_flow': investing_cf,
            'financing_cash_flow': financing_cf,
            'net_cash_flow': net_cf,
            'cash_position': cash_position,
            'forecast': forecast
        }
    
    def _calculate_operating_cash_flow(self, start_date: date, end_date: date) -> Decimal:
        """Calculate operating cash flow."""
        
        # Get customer payments (cash inflows)
        customer_payments = Payment.objects.filter(
            company=self.company,
            payment_type='CUSTOMER',
            state='CONFIRMED',
            payment_date__range=[start_date, end_date]
        ).aggregate(
            total=Coalesce(Sum('amount'), Decimal('0'))
        )['total']
        
        # Get supplier payments (cash outflows)
        supplier_payments = Payment.objects.filter(
            company=self.company,
            payment_type='SUPPLIER',
            state='CONFIRMED',
            payment_date__range=[start_date, end_date]
        ).aggregate(
            total=Coalesce(Sum('amount'), Decimal('0'))
        )['total']
        
        return customer_payments - supplier_payments
    
    def _calculate_investing_cash_flow(self, start_date: date, end_date: date) -> Decimal:
        """Calculate investing cash flow."""
        
        # Get asset purchases and sales
        asset_accounts = Account.objects.filter(
            company=self.company,
            code__startswith='2',  # Fixed assets
            is_active=True
        )
        
        investing_cf = Decimal('0')
        for account in asset_accounts:
            lines = JournalEntryLine.objects.filter(
                account=account,
                journal_entry__date__range=[start_date, end_date],
                journal_entry__state='POSTED'
            )
            
            # Asset purchases (debits) are cash outflows
            purchases = lines.aggregate(
                total=Coalesce(Sum('debit_amount'), Decimal('0'))
            )['total']
            
            # Asset sales (credits) are cash inflows
            sales = lines.aggregate(
                total=Coalesce(Sum('credit_amount'), Decimal('0'))
            )['total']
            
            investing_cf += sales - purchases
        
        return investing_cf
    
    def _calculate_financing_cash_flow(self, start_date: date, end_date: date) -> Decimal:
        """Calculate financing cash flow."""
        
        # Get equity and debt accounts
        equity_accounts = Account.objects.filter(
            company=self.company,
            code__startswith='1',  # Equity accounts
            is_active=True
        )
        
        debt_accounts = Account.objects.filter(
            company=self.company,
            code__startswith='4',  # Liability accounts
            is_active=True
        )
        
        financing_cf = Decimal('0')
        
        # Process equity changes
        for account in equity_accounts:
            lines = JournalEntryLine.objects.filter(
                account=account,
                journal_entry__date__range=[start_date, end_date],
                journal_entry__state='POSTED'
            )
            
            credits = lines.aggregate(
                total=Coalesce(Sum('credit_amount'), Decimal('0'))
            )['total']
            
            debits = lines.aggregate(
                total=Coalesce(Sum('debit_amount'), Decimal('0'))
            )['total']
            
            financing_cf += credits - debits
        
        # Process debt changes
        for account in debt_accounts:
            lines = JournalEntryLine.objects.filter(
                account=account,
                journal_entry__date__range=[start_date, end_date],
                journal_entry__state='POSTED'
            )
            
            credits = lines.aggregate(
                total=Coalesce(Sum('credit_amount'), Decimal('0'))
            )['total']
            
            debits = lines.aggregate(
                total=Coalesce(Sum('debit_amount'), Decimal('0'))
            )['total']
            
            financing_cf += credits - debits
        
        return financing_cf
    
    def _get_cash_position(self, cash_accounts, as_of_date: date) -> Dict[str, Any]:
        """Get current cash position."""
        
        total_cash = Decimal('0')
        account_balances = {}
        
        for account in cash_accounts:
            balance = account.current_balance
            total_cash += balance
            account_balances[account.code] = {
                'name': account.name,
                'balance': balance
            }
        
        return {
            'total': total_cash,
            'accounts': account_balances,
            'as_of_date': as_of_date
        }
    
    def _forecast_cash_flow(self, days: int) -> Dict[str, Any]:
        """Forecast cash flow for the next N days."""
        
        # Simple forecast based on historical patterns
        # TODO: Implement more sophisticated forecasting algorithms
        
        today = date.today()
        forecast_date = today + timedelta(days=days)
        
        # Get average daily cash flow from last 30 days
        start_date = today - timedelta(days=30)
        historical_cf = self.generate_cash_flow_analysis(start_date, today)
        
        daily_avg = historical_cf['net_cash_flow'] / 30
        forecasted_cf = daily_avg * days
        
        return {
            'forecast_date': forecast_date,
            'forecasted_cash_flow': forecasted_cf,
            'daily_average': daily_avg,
            'method': 'Historical Average'
        }
    
    def generate_balance_sheet_comparison(
        self,
        as_of_date: date,
        comparison_date: date
    ) -> Dict[str, Any]:
        """Generate balance sheet with period comparison."""
        
        current_bs = self._get_balance_sheet_data(as_of_date)
        comparison_bs = self._get_balance_sheet_data(comparison_date)
        
        # Calculate changes
        changes = self._calculate_balance_sheet_changes(current_bs, comparison_bs)
        
        return {
            'current': {
                'date': as_of_date,
                'data': current_bs
            },
            'comparison': {
                'date': comparison_date,
                'data': comparison_bs
            },
            'changes': changes
        }
    
    def _get_balance_sheet_data(self, as_of_date: date) -> Dict[str, Any]:
        """Get balance sheet data as of a specific date."""
        
        # Assets (accounts starting with 2, 3, 5)
        asset_accounts = Account.objects.filter(
            company=self.company,
            code__regex=r'^[235]',
            is_active=True
        )
        
        # Liabilities (accounts starting with 4)
        liability_accounts = Account.objects.filter(
            company=self.company,
            code__startswith='4',
            is_active=True
        )
        
        # Equity (accounts starting with 1)
        equity_accounts = Account.objects.filter(
            company=self.company,
            code__startswith='1',
            is_active=True
        )
        
        assets = self._get_account_group_balances(asset_accounts, as_of_date)
        liabilities = self._get_account_group_balances(liability_accounts, as_of_date)
        equity = self._get_account_group_balances(equity_accounts, as_of_date)
        
        total_assets = sum(item['balance'] for item in assets.values())
        total_liabilities = sum(item['balance'] for item in liabilities.values())
        total_equity = sum(item['balance'] for item in equity.values())
        
        return {
            'assets': {
                'accounts': assets,
                'total': total_assets
            },
            'liabilities': {
                'accounts': liabilities,
                'total': total_liabilities
            },
            'equity': {
                'accounts': equity,
                'total': total_equity
            },
            'total_liabilities_equity': total_liabilities + total_equity
        }
    
    def _get_account_group_balances(self, accounts, as_of_date: date) -> Dict[str, Dict[str, Any]]:
        """Get balances for a group of accounts as of a specific date."""
        
        result = {}
        
        for account in accounts:
            # Calculate balance as of date
            lines = JournalEntryLine.objects.filter(
                account=account,
                journal_entry__date__lte=as_of_date,
                journal_entry__state='POSTED'
            )
            
            debits = lines.aggregate(
                total=Coalesce(Sum('debit_amount'), Decimal('0'))
            )['total']
            
            credits = lines.aggregate(
                total=Coalesce(Sum('credit_amount'), Decimal('0'))
            )['total']
            
            # Calculate balance based on account normal balance
            if account.account_type.normal_balance == 'DEBIT':
                balance = debits - credits
            else:
                balance = credits - debits
            
            if balance != 0:
                result[account.code] = {
                    'name': account.name,
                    'balance': balance,
                    'account_type': account.account_type.name
                }
        
        return result
    
    def _calculate_balance_sheet_changes(
        self,
        current_bs: Dict[str, Any],
        comparison_bs: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Calculate changes between two balance sheets."""
        
        changes = {}
        
        # Asset changes
        current_assets = current_bs['assets']['total']
        comparison_assets = comparison_bs['assets']['total']
        asset_change = current_assets - comparison_assets
        
        changes['assets'] = {
            'absolute': asset_change,
            'percentage': (asset_change / comparison_assets * 100) if comparison_assets > 0 else 0
        }
        
        # Liability changes
        current_liabilities = current_bs['liabilities']['total']
        comparison_liabilities = comparison_bs['liabilities']['total']
        liability_change = current_liabilities - comparison_liabilities
        
        changes['liabilities'] = {
            'absolute': liability_change,
            'percentage': (liability_change / comparison_liabilities * 100) if comparison_liabilities > 0 else 0
        }
        
        # Equity changes
        current_equity = current_bs['equity']['total']
        comparison_equity = comparison_bs['equity']['total']
        equity_change = current_equity - comparison_equity
        
        changes['equity'] = {
            'absolute': equity_change,
            'percentage': (equity_change / comparison_equity * 100) if comparison_equity > 0 else 0
        }
        
        return changes


class KPIAnalyticsService:
    """Service for generating KPI and analytics reports."""

    def __init__(self, company: Company):
        self.company = company

    def generate_financial_kpis(
        self,
        start_date: date,
        end_date: date
    ) -> Dict[str, Any]:
        """Generate key financial KPIs."""

        # Revenue metrics
        total_revenue = self._get_total_revenue(start_date, end_date)
        revenue_growth = self._calculate_revenue_growth(start_date, end_date)

        # Profitability metrics
        gross_profit = self._get_gross_profit(start_date, end_date)
        net_profit = self._get_net_profit(start_date, end_date)
        gross_margin = (gross_profit / total_revenue * 100) if total_revenue > 0 else 0
        net_margin = (net_profit / total_revenue * 100) if total_revenue > 0 else 0

        # Liquidity metrics
        current_ratio = self._calculate_current_ratio()
        quick_ratio = self._calculate_quick_ratio()
        cash_ratio = self._calculate_cash_ratio()

        # Efficiency metrics
        asset_turnover = self._calculate_asset_turnover(start_date, end_date)
        inventory_turnover = self._calculate_inventory_turnover(start_date, end_date)
        receivables_turnover = self._calculate_receivables_turnover(start_date, end_date)

        return {
            'period': {
                'start_date': start_date,
                'end_date': end_date
            },
            'revenue_metrics': {
                'total_revenue': total_revenue,
                'revenue_growth': revenue_growth
            },
            'profitability_metrics': {
                'gross_profit': gross_profit,
                'net_profit': net_profit,
                'gross_margin': gross_margin,
                'net_margin': net_margin
            },
            'liquidity_metrics': {
                'current_ratio': current_ratio,
                'quick_ratio': quick_ratio,
                'cash_ratio': cash_ratio
            },
            'efficiency_metrics': {
                'asset_turnover': asset_turnover,
                'inventory_turnover': inventory_turnover,
                'receivables_turnover': receivables_turnover
            }
        }

    def _get_total_revenue(self, start_date: date, end_date: date) -> Decimal:
        """Get total revenue for period."""
        return Invoice.objects.filter(
            company=self.company,
            invoice_type='CUSTOMER',
            state='POSTED',
            invoice_date__range=[start_date, end_date]
        ).aggregate(
            total=Coalesce(Sum('total_amount'), Decimal('0'))
        )['total']

    def _calculate_revenue_growth(self, start_date: date, end_date: date) -> Decimal:
        """Calculate revenue growth compared to previous period."""
        period_days = (end_date - start_date).days
        prev_start = start_date - timedelta(days=period_days)
        prev_end = start_date - timedelta(days=1)

        current_revenue = self._get_total_revenue(start_date, end_date)
        previous_revenue = self._get_total_revenue(prev_start, prev_end)

        if previous_revenue > 0:
            return (current_revenue - previous_revenue) / previous_revenue * 100
        return Decimal('0')

    def _get_gross_profit(self, start_date: date, end_date: date) -> Decimal:
        """Get gross profit for period."""
        # Revenue - Cost of Goods Sold
        revenue = self._get_total_revenue(start_date, end_date)

        # Get COGS from expense accounts (typically 6xxx accounts)
        cogs_accounts = Account.objects.filter(
            company=self.company,
            code__startswith='6',
            is_active=True
        )

        cogs = Decimal('0')
        for account in cogs_accounts:
            lines = JournalEntryLine.objects.filter(
                account=account,
                journal_entry__date__range=[start_date, end_date],
                journal_entry__state='POSTED'
            )
            cogs += lines.aggregate(
                total=Coalesce(Sum('debit_amount'), Decimal('0'))
            )['total']

        return revenue - cogs

    def _get_net_profit(self, start_date: date, end_date: date) -> Decimal:
        """Get net profit for period."""
        # Get all revenue
        revenue_accounts = Account.objects.filter(
            company=self.company,
            code__startswith='7',
            is_active=True
        )

        total_revenue = Decimal('0')
        for account in revenue_accounts:
            lines = JournalEntryLine.objects.filter(
                account=account,
                journal_entry__date__range=[start_date, end_date],
                journal_entry__state='POSTED'
            )
            total_revenue += lines.aggregate(
                total=Coalesce(Sum('credit_amount'), Decimal('0'))
            )['total']

        # Get all expenses
        expense_accounts = Account.objects.filter(
            company=self.company,
            code__startswith='6',
            is_active=True
        )

        total_expenses = Decimal('0')
        for account in expense_accounts:
            lines = JournalEntryLine.objects.filter(
                account=account,
                journal_entry__date__range=[start_date, end_date],
                journal_entry__state='POSTED'
            )
            total_expenses += lines.aggregate(
                total=Coalesce(Sum('debit_amount'), Decimal('0'))
            )['total']

        return total_revenue - total_expenses

    def _calculate_current_ratio(self) -> Decimal:
        """Calculate current ratio (current assets / current liabilities)."""
        # Current assets (typically 3xxx and 5xxx accounts)
        current_assets = Account.objects.filter(
            company=self.company,
            code__regex=r'^[35]',
            is_active=True
        ).aggregate(
            total=Coalesce(Sum('current_balance'), Decimal('0'))
        )['total']

        # Current liabilities (typically 44xx accounts)
        current_liabilities = Account.objects.filter(
            company=self.company,
            code__startswith='44',
            is_active=True
        ).aggregate(
            total=Coalesce(Sum('current_balance'), Decimal('0'))
        )['total']

        if current_liabilities > 0:
            return current_assets / current_liabilities
        return Decimal('0')

    def _calculate_quick_ratio(self) -> Decimal:
        """Calculate quick ratio ((current assets - inventory) / current liabilities)."""
        # Current assets minus inventory
        current_assets = Account.objects.filter(
            company=self.company,
            code__regex=r'^[35]',
            is_active=True
        ).aggregate(
            total=Coalesce(Sum('current_balance'), Decimal('0'))
        )['total']

        # Inventory (typically 31xx accounts)
        inventory = Account.objects.filter(
            company=self.company,
            code__startswith='31',
            is_active=True
        ).aggregate(
            total=Coalesce(Sum('current_balance'), Decimal('0'))
        )['total']

        # Current liabilities
        current_liabilities = Account.objects.filter(
            company=self.company,
            code__startswith='44',
            is_active=True
        ).aggregate(
            total=Coalesce(Sum('current_balance'), Decimal('0'))
        )['total']

        quick_assets = current_assets - inventory

        if current_liabilities > 0:
            return quick_assets / current_liabilities
        return Decimal('0')

    def _calculate_cash_ratio(self) -> Decimal:
        """Calculate cash ratio (cash / current liabilities)."""
        # Cash and cash equivalents (511x accounts)
        cash = Account.objects.filter(
            company=self.company,
            code__startswith='511',
            is_active=True
        ).aggregate(
            total=Coalesce(Sum('current_balance'), Decimal('0'))
        )['total']

        # Current liabilities
        current_liabilities = Account.objects.filter(
            company=self.company,
            code__startswith='44',
            is_active=True
        ).aggregate(
            total=Coalesce(Sum('current_balance'), Decimal('0'))
        )['total']

        if current_liabilities > 0:
            return cash / current_liabilities
        return Decimal('0')

    def _calculate_asset_turnover(self, start_date: date, end_date: date) -> Decimal:
        """Calculate asset turnover ratio (revenue / average total assets)."""
        revenue = self._get_total_revenue(start_date, end_date)

        # Get total assets
        total_assets = Account.objects.filter(
            company=self.company,
            code__regex=r'^[235]',
            is_active=True
        ).aggregate(
            total=Coalesce(Sum('current_balance'), Decimal('0'))
        )['total']

        if total_assets > 0:
            return revenue / total_assets
        return Decimal('0')

    def _calculate_inventory_turnover(self, start_date: date, end_date: date) -> Decimal:
        """Calculate inventory turnover ratio (COGS / average inventory)."""
        # Get COGS
        cogs_accounts = Account.objects.filter(
            company=self.company,
            code__startswith='6',
            is_active=True
        )

        cogs = Decimal('0')
        for account in cogs_accounts:
            lines = JournalEntryLine.objects.filter(
                account=account,
                journal_entry__date__range=[start_date, end_date],
                journal_entry__state='POSTED'
            )
            cogs += lines.aggregate(
                total=Coalesce(Sum('debit_amount'), Decimal('0'))
            )['total']

        # Get average inventory
        inventory = Account.objects.filter(
            company=self.company,
            code__startswith='31',
            is_active=True
        ).aggregate(
            total=Coalesce(Sum('current_balance'), Decimal('0'))
        )['total']

        if inventory > 0:
            return cogs / inventory
        return Decimal('0')

    def _calculate_receivables_turnover(self, start_date: date, end_date: date) -> Decimal:
        """Calculate receivables turnover ratio (revenue / average receivables)."""
        revenue = self._get_total_revenue(start_date, end_date)

        # Get accounts receivable (3421 - Clients)
        receivables = Account.objects.filter(
            company=self.company,
            code='3421',
            is_active=True
        ).aggregate(
            total=Coalesce(Sum('current_balance'), Decimal('0'))
        )['total']

        if receivables > 0:
            return revenue / receivables
        return Decimal('0')


class AgedReportsService:
    """Service for generating aged AR/AP reports."""

    def __init__(self, company: Company):
        self.company = company

    def generate_aged_ar_report(self, as_of_date: date = None) -> Dict[str, Any]:
        """
        Generate Aged Accounts Receivable report.

        Args:
            as_of_date: Date to calculate aging as of (defaults to today)

        Returns:
            Dictionary with aged AR data
        """
        if as_of_date is None:
            as_of_date = timezone.now().date()

        from modules.invoicing.models import Invoice

        # Get all unpaid customer invoices
        unpaid_invoices = Invoice.objects.filter(
            company=self.company,
            invoice_type='CUSTOMER',
            status__in=['APPROVED', 'SENT'],
            due_date__lte=as_of_date
        ).select_related('customer')

        aged_data = {
            'as_of_date': as_of_date,
            'summary': {
                'current': Decimal('0'),
                '1_30_days': Decimal('0'),
                '31_60_days': Decimal('0'),
                '61_90_days': Decimal('0'),
                'over_90_days': Decimal('0'),
                'total': Decimal('0')
            },
            'customers': []
        }

        customer_data = {}

        for invoice in unpaid_invoices:
            days_overdue = (as_of_date - invoice.due_date).days
            outstanding_amount = invoice.total_amount - (invoice.paid_amount or Decimal('0'))

            if outstanding_amount <= 0:
                continue

            # Categorize by aging bucket
            if days_overdue <= 0:
                bucket = 'current'
            elif days_overdue <= 30:
                bucket = '1_30_days'
            elif days_overdue <= 60:
                bucket = '31_60_days'
            elif days_overdue <= 90:
                bucket = '61_90_days'
            else:
                bucket = 'over_90_days'

            # Update summary
            aged_data['summary'][bucket] += outstanding_amount
            aged_data['summary']['total'] += outstanding_amount

            # Update customer data
            customer_key = str(invoice.customer.id) if invoice.customer else 'unknown'
            if customer_key not in customer_data:
                customer_data[customer_key] = {
                    'customer_id': customer_key,
                    'customer_name': invoice.customer.name if invoice.customer else 'Unknown',
                    'current': Decimal('0'),
                    '1_30_days': Decimal('0'),
                    '31_60_days': Decimal('0'),
                    '61_90_days': Decimal('0'),
                    'over_90_days': Decimal('0'),
                    'total': Decimal('0'),
                    'invoices': []
                }

            customer_data[customer_key][bucket] += outstanding_amount
            customer_data[customer_key]['total'] += outstanding_amount
            customer_data[customer_key]['invoices'].append({
                'invoice_number': invoice.sequence_no,
                'invoice_date': invoice.issue_date,
                'due_date': invoice.due_date,
                'days_overdue': max(0, days_overdue),
                'outstanding_amount': outstanding_amount
            })

        aged_data['customers'] = list(customer_data.values())

        return aged_data

    def generate_aged_ap_report(self, as_of_date: date = None) -> Dict[str, Any]:
        """
        Generate Aged Accounts Payable report.

        Args:
            as_of_date: Date to calculate aging as of (defaults to today)

        Returns:
            Dictionary with aged AP data
        """
        if as_of_date is None:
            as_of_date = timezone.now().date()

        from modules.invoicing.models import Invoice

        # Get all unpaid supplier invoices
        unpaid_invoices = Invoice.objects.filter(
            company=self.company,
            invoice_type='SUPPLIER',
            status__in=['APPROVED', 'SENT'],
            due_date__lte=as_of_date
        ).select_related('supplier')

        aged_data = {
            'as_of_date': as_of_date,
            'summary': {
                'current': Decimal('0'),
                '1_30_days': Decimal('0'),
                '31_60_days': Decimal('0'),
                '61_90_days': Decimal('0'),
                'over_90_days': Decimal('0'),
                'total': Decimal('0')
            },
            'suppliers': []
        }

        supplier_data = {}

        for invoice in unpaid_invoices:
            days_overdue = (as_of_date - invoice.due_date).days
            outstanding_amount = invoice.total_amount - (invoice.paid_amount or Decimal('0'))

            if outstanding_amount <= 0:
                continue

            # Categorize by aging bucket
            if days_overdue <= 0:
                bucket = 'current'
            elif days_overdue <= 30:
                bucket = '1_30_days'
            elif days_overdue <= 60:
                bucket = '31_60_days'
            elif days_overdue <= 90:
                bucket = '61_90_days'
            else:
                bucket = 'over_90_days'

            # Update summary
            aged_data['summary'][bucket] += outstanding_amount
            aged_data['summary']['total'] += outstanding_amount

            # Update supplier data
            supplier_key = str(invoice.supplier.id) if invoice.supplier else 'unknown'
            if supplier_key not in supplier_data:
                supplier_data[supplier_key] = {
                    'supplier_id': supplier_key,
                    'supplier_name': invoice.supplier.name if invoice.supplier else 'Unknown',
                    'current': Decimal('0'),
                    '1_30_days': Decimal('0'),
                    '31_60_days': Decimal('0'),
                    '61_90_days': Decimal('0'),
                    'over_90_days': Decimal('0'),
                    'total': Decimal('0'),
                    'invoices': []
                }

            supplier_data[supplier_key][bucket] += outstanding_amount
            supplier_data[supplier_key]['total'] += outstanding_amount
            supplier_data[supplier_key]['invoices'].append({
                'invoice_number': invoice.sequence_no,
                'invoice_date': invoice.issue_date,
                'due_date': invoice.due_date,
                'days_overdue': max(0, days_overdue),
                'outstanding_amount': outstanding_amount
            })

        aged_data['suppliers'] = list(supplier_data.values())

        return aged_data


class StockValuationReportService:
    """Service for generating stock valuation reports."""

    def __init__(self, company: Company):
        self.company = company

    def generate_stock_valuation_report(self, as_of_date: date = None, warehouse_id: str = None) -> Dict[str, Any]:
        """
        Generate Stock Valuation report by warehouse/product.

        Args:
            as_of_date: Date to calculate valuation as of (defaults to today)
            warehouse_id: Optional warehouse filter

        Returns:
            Dictionary with stock valuation data
        """
        if as_of_date is None:
            as_of_date = timezone.now().date()

        from modules.inventory.models import StockMove, Warehouse
        from modules.catalog.models import Product

        # Get stock movements up to the as_of_date
        stock_moves = StockMove.objects.filter(
            company=self.company,
            created_at__date__lte=as_of_date
        ).select_related('product', 'warehouse')

        if warehouse_id:
            stock_moves = stock_moves.filter(warehouse_id=warehouse_id)

        valuation_data = {
            'as_of_date': as_of_date,
            'warehouse_filter': warehouse_id,
            'summary': {
                'total_quantity': Decimal('0'),
                'total_value': Decimal('0'),
                'total_products': 0,
                'total_warehouses': 0
            },
            'warehouses': []
        }

        warehouse_data = {}

        # Calculate stock levels and values by warehouse and product
        for move in stock_moves:
            warehouse_key = str(move.warehouse.id) if move.warehouse else 'unknown'
            product_key = str(move.product.id)

            if warehouse_key not in warehouse_data:
                warehouse_data[warehouse_key] = {
                    'warehouse_id': warehouse_key,
                    'warehouse_name': move.warehouse.name if move.warehouse else 'Unknown',
                    'total_quantity': Decimal('0'),
                    'total_value': Decimal('0'),
                    'products': {}
                }

            if product_key not in warehouse_data[warehouse_key]['products']:
                warehouse_data[warehouse_key]['products'][product_key] = {
                    'product_id': product_key,
                    'product_sku': move.product.sku,
                    'product_name': move.product.name,
                    'uom': move.product.uom.code if move.product.uom else '',
                    'quantity': Decimal('0'),
                    'average_cost': Decimal('0'),
                    'total_value': Decimal('0'),
                    'movements': []
                }

            product_data = warehouse_data[warehouse_key]['products'][product_key]

            # Update quantities based on move type
            if move.move_type in ['PO_RECEIPT', 'TRANSFER_IN', 'ADJUSTMENT_IN']:
                product_data['quantity'] += move.quantity
                if move.unit_cost:
                    # Weighted average cost calculation
                    current_value = product_data['total_value']
                    new_value = move.quantity * move.unit_cost
                    total_quantity = product_data['quantity']

                    if total_quantity > 0:
                        product_data['average_cost'] = (current_value + new_value) / total_quantity
                        product_data['total_value'] = current_value + new_value
            elif move.move_type in ['SO_DELIVERY', 'TRANSFER_OUT', 'ADJUSTMENT_OUT']:
                product_data['quantity'] -= move.quantity
                # Reduce value proportionally
                if product_data['quantity'] > 0:
                    product_data['total_value'] = product_data['quantity'] * product_data['average_cost']
                else:
                    product_data['total_value'] = Decimal('0')
                    product_data['average_cost'] = Decimal('0')

            product_data['movements'].append({
                'date': move.created_at.date(),
                'move_type': move.move_type,
                'quantity': move.quantity,
                'unit_cost': move.unit_cost or Decimal('0'),
                'reference': move.reference or ''
            })

        # Convert to list format and calculate totals
        warehouses_list = []
        total_products = set()

        for warehouse_key, warehouse_info in warehouse_data.items():
            warehouse_summary = {
                'warehouse_id': warehouse_info['warehouse_id'],
                'warehouse_name': warehouse_info['warehouse_name'],
                'total_quantity': Decimal('0'),
                'total_value': Decimal('0'),
                'product_count': 0,
                'products': []
            }

            for product_key, product_info in warehouse_info['products'].items():
                if product_info['quantity'] > 0:  # Only include products with positive stock
                    warehouse_summary['products'].append(product_info)
                    warehouse_summary['total_quantity'] += product_info['quantity']
                    warehouse_summary['total_value'] += product_info['total_value']
                    warehouse_summary['product_count'] += 1
                    total_products.add(product_key)

            if warehouse_summary['product_count'] > 0:
                warehouses_list.append(warehouse_summary)
                valuation_data['summary']['total_quantity'] += warehouse_summary['total_quantity']
                valuation_data['summary']['total_value'] += warehouse_summary['total_value']

        valuation_data['warehouses'] = warehouses_list
        valuation_data['summary']['total_products'] = len(total_products)
        valuation_data['summary']['total_warehouses'] = len(warehouses_list)

        return valuation_data

    def generate_stock_movement_report(self, start_date: date, end_date: date, product_id: str = None) -> Dict[str, Any]:
        """
        Generate Stock Movement report for a period.

        Args:
            start_date: Start date for the report
            end_date: End date for the report
            product_id: Optional product filter

        Returns:
            Dictionary with stock movement data
        """
        from modules.inventory.models import StockMove

        stock_moves = StockMove.objects.filter(
            company=self.company,
            created_at__date__range=[start_date, end_date]
        ).select_related('product', 'warehouse').order_by('created_at')

        if product_id:
            stock_moves = stock_moves.filter(product_id=product_id)

        movement_data = {
            'period': {
                'start_date': start_date,
                'end_date': end_date
            },
            'product_filter': product_id,
            'summary': {
                'total_movements': stock_moves.count(),
                'total_inbound': Decimal('0'),
                'total_outbound': Decimal('0'),
                'net_movement': Decimal('0')
            },
            'movements': []
        }

        for move in stock_moves:
            movement_info = {
                'date': move.created_at.date(),
                'product_sku': move.product.sku,
                'product_name': move.product.name,
                'warehouse': move.warehouse.name if move.warehouse else 'Unknown',
                'move_type': move.move_type,
                'quantity': move.quantity,
                'unit_cost': move.unit_cost or Decimal('0'),
                'total_value': (move.quantity * (move.unit_cost or Decimal('0'))),
                'reference': move.reference or ''
            }

            movement_data['movements'].append(movement_info)

            # Update summary
            if move.move_type in ['PO_RECEIPT', 'TRANSFER_IN', 'ADJUSTMENT_IN']:
                movement_data['summary']['total_inbound'] += move.quantity
            elif move.move_type in ['SO_DELIVERY', 'TRANSFER_OUT', 'ADJUSTMENT_OUT']:
                movement_data['summary']['total_outbound'] += move.quantity

        movement_data['summary']['net_movement'] = (
            movement_data['summary']['total_inbound'] -
            movement_data['summary']['total_outbound']
        )

        return movement_data


class VATReportService:
    """Service for generating VAT-related reports."""

    def __init__(self, company: Company):
        self.company = company

    def generate_sales_by_vat_rate_report(self, start_date: date, end_date: date) -> Dict[str, Any]:
        """
        Generate Sales by VAT Rate report.

        Args:
            start_date: Start date for the report
            end_date: End date for the report

        Returns:
            Dictionary with sales by VAT rate data
        """
        from modules.invoicing.models import Invoice, InvoiceLine

        # Get all customer invoices in the period
        invoices = Invoice.objects.filter(
            company=self.company,
            invoice_type='CUSTOMER',
            status__in=['APPROVED', 'SENT', 'PAID'],
            issue_date__range=[start_date, end_date]
        ).prefetch_related('lines')

        vat_data = {
            'period': {
                'start_date': start_date,
                'end_date': end_date
            },
            'summary': {
                'total_base_amount': Decimal('0'),
                'total_vat_amount': Decimal('0'),
                'total_amount': Decimal('0'),
                'invoice_count': invoices.count()
            },
            'vat_rates': []
        }

        vat_rate_data = {}

        for invoice in invoices:
            for line in invoice.lines.all():
                # Parse tax information from line
                taxes_applied = line.taxes_applied or []

                for tax_info in taxes_applied:
                    rate_key = f"{tax_info.get('rate', 0):.2f}%"

                    if rate_key not in vat_rate_data:
                        vat_rate_data[rate_key] = {
                            'vat_rate': tax_info.get('rate', 0),
                            'base_amount': Decimal('0'),
                            'vat_amount': Decimal('0'),
                            'total_amount': Decimal('0'),
                            'transaction_count': 0,
                            'invoices': []
                        }

                    line_base = line.quantity * line.unit_price * (1 - (line.discount_percent or 0) / 100)
                    line_vat = tax_info.get('amount', 0)
                    line_total = line_base + line_vat

                    vat_rate_data[rate_key]['base_amount'] += line_base
                    vat_rate_data[rate_key]['vat_amount'] += Decimal(str(line_vat))
                    vat_rate_data[rate_key]['total_amount'] += line_total
                    vat_rate_data[rate_key]['transaction_count'] += 1

                    # Add invoice reference if not already added
                    invoice_ref = {
                        'invoice_number': invoice.sequence_no,
                        'invoice_date': invoice.issue_date,
                        'customer_name': invoice.customer.name if invoice.customer else 'Unknown',
                        'base_amount': line_base,
                        'vat_amount': Decimal(str(line_vat)),
                        'total_amount': line_total
                    }

                    if invoice_ref not in vat_rate_data[rate_key]['invoices']:
                        vat_rate_data[rate_key]['invoices'].append(invoice_ref)

        # Convert to list and calculate totals
        for rate_info in vat_rate_data.values():
            vat_data['vat_rates'].append(rate_info)
            vat_data['summary']['total_base_amount'] += rate_info['base_amount']
            vat_data['summary']['total_vat_amount'] += rate_info['vat_amount']
            vat_data['summary']['total_amount'] += rate_info['total_amount']

        # Sort by VAT rate
        vat_data['vat_rates'].sort(key=lambda x: x['vat_rate'])

        return vat_data

    def generate_purchases_by_vat_rate_report(self, start_date: date, end_date: date) -> Dict[str, Any]:
        """
        Generate Purchases by VAT Rate report.

        Args:
            start_date: Start date for the report
            end_date: End date for the report

        Returns:
            Dictionary with purchases by VAT rate data
        """
        from modules.invoicing.models import Invoice, InvoiceLine

        # Get all supplier invoices in the period
        invoices = Invoice.objects.filter(
            company=self.company,
            invoice_type='SUPPLIER',
            status__in=['APPROVED', 'SENT', 'PAID'],
            issue_date__range=[start_date, end_date]
        ).prefetch_related('lines')

        vat_data = {
            'period': {
                'start_date': start_date,
                'end_date': end_date
            },
            'summary': {
                'total_base_amount': Decimal('0'),
                'total_vat_amount': Decimal('0'),
                'total_withheld_vat': Decimal('0'),
                'total_amount': Decimal('0'),
                'invoice_count': invoices.count()
            },
            'vat_rates': []
        }

        vat_rate_data = {}

        for invoice in invoices:
            for line in invoice.lines.all():
                # Parse tax information from line
                taxes_applied = line.taxes_applied or []

                for tax_info in taxes_applied:
                    rate_key = f"{tax_info.get('rate', 0):.2f}%"

                    if rate_key not in vat_rate_data:
                        vat_rate_data[rate_key] = {
                            'vat_rate': tax_info.get('rate', 0),
                            'base_amount': Decimal('0'),
                            'vat_amount': Decimal('0'),
                            'withheld_vat_amount': Decimal('0'),
                            'total_amount': Decimal('0'),
                            'transaction_count': 0,
                            'invoices': []
                        }

                    line_base = line.quantity * line.unit_price * (1 - (line.discount_percent or 0) / 100)
                    line_vat = tax_info.get('amount', 0)
                    line_withheld = tax_info.get('withheld_amount', 0)
                    line_total = line_base + line_vat - line_withheld

                    vat_rate_data[rate_key]['base_amount'] += line_base
                    vat_rate_data[rate_key]['vat_amount'] += Decimal(str(line_vat))
                    vat_rate_data[rate_key]['withheld_vat_amount'] += Decimal(str(line_withheld))
                    vat_rate_data[rate_key]['total_amount'] += line_total
                    vat_rate_data[rate_key]['transaction_count'] += 1

                    # Add invoice reference if not already added
                    invoice_ref = {
                        'invoice_number': invoice.sequence_no,
                        'invoice_date': invoice.issue_date,
                        'supplier_name': invoice.supplier.name if invoice.supplier else 'Unknown',
                        'base_amount': line_base,
                        'vat_amount': Decimal(str(line_vat)),
                        'withheld_vat_amount': Decimal(str(line_withheld)),
                        'total_amount': line_total
                    }

                    if invoice_ref not in vat_rate_data[rate_key]['invoices']:
                        vat_rate_data[rate_key]['invoices'].append(invoice_ref)

        # Convert to list and calculate totals
        for rate_info in vat_rate_data.values():
            vat_data['vat_rates'].append(rate_info)
            vat_data['summary']['total_base_amount'] += rate_info['base_amount']
            vat_data['summary']['total_vat_amount'] += rate_info['vat_amount']
            vat_data['summary']['total_withheld_vat'] += rate_info['withheld_vat_amount']
            vat_data['summary']['total_amount'] += rate_info['total_amount']

        # Sort by VAT rate
        vat_data['vat_rates'].sort(key=lambda x: x['vat_rate'])

        return vat_data


class ReportExportService:
    """Service for exporting reports to CSV/XLSX with French headers."""

    def __init__(self, company: Company):
        self.company = company

    def export_aged_ar_to_csv(self, aged_data: Dict[str, Any]) -> str:
        """Export Aged AR report to CSV with French headers."""
        import csv
        import io

        output = io.StringIO()
        writer = csv.writer(output)

        # French headers for Aged AR
        writer.writerow([
            'Client',
            'Courant (MAD)',
            '1-30 jours (MAD)',
            '31-60 jours (MAD)',
            '61-90 jours (MAD)',
            'Plus de 90 jours (MAD)',
            'Total (MAD)'
        ])

        # Write customer data
        for customer in aged_data['customers']:
            writer.writerow([
                customer['customer_name'],
                f"{customer['current']:.2f}",
                f"{customer['1_30_days']:.2f}",
                f"{customer['31_60_days']:.2f}",
                f"{customer['61_90_days']:.2f}",
                f"{customer['over_90_days']:.2f}",
                f"{customer['total']:.2f}"
            ])

        # Write summary
        writer.writerow([])
        writer.writerow(['TOTAL GÉNÉRAL'] + [
            f"{aged_data['summary']['current']:.2f}",
            f"{aged_data['summary']['1_30_days']:.2f}",
            f"{aged_data['summary']['31_60_days']:.2f}",
            f"{aged_data['summary']['61_90_days']:.2f}",
            f"{aged_data['summary']['over_90_days']:.2f}",
            f"{aged_data['summary']['total']:.2f}"
        ])

        return output.getvalue()

    def export_aged_ap_to_csv(self, aged_data: Dict[str, Any]) -> str:
        """Export Aged AP report to CSV with French headers."""
        import csv
        import io

        output = io.StringIO()
        writer = csv.writer(output)

        # French headers for Aged AP
        writer.writerow([
            'Fournisseur',
            'Courant (MAD)',
            '1-30 jours (MAD)',
            '31-60 jours (MAD)',
            '61-90 jours (MAD)',
            'Plus de 90 jours (MAD)',
            'Total (MAD)'
        ])

        # Write supplier data
        for supplier in aged_data['suppliers']:
            writer.writerow([
                supplier['supplier_name'],
                f"{supplier['current']:.2f}",
                f"{supplier['1_30_days']:.2f}",
                f"{supplier['31_60_days']:.2f}",
                f"{supplier['61_90_days']:.2f}",
                f"{supplier['over_90_days']:.2f}",
                f"{supplier['total']:.2f}"
            ])

        # Write summary
        writer.writerow([])
        writer.writerow(['TOTAL GÉNÉRAL'] + [
            f"{aged_data['summary']['current']:.2f}",
            f"{aged_data['summary']['1_30_days']:.2f}",
            f"{aged_data['summary']['31_60_days']:.2f}",
            f"{aged_data['summary']['61_90_days']:.2f}",
            f"{aged_data['summary']['over_90_days']:.2f}",
            f"{aged_data['summary']['total']:.2f}"
        ])

        return output.getvalue()

    def export_stock_valuation_to_csv(self, valuation_data: Dict[str, Any]) -> str:
        """Export Stock Valuation report to CSV with French headers."""
        import csv
        import io

        output = io.StringIO()
        writer = csv.writer(output)

        # French headers for Stock Valuation
        writer.writerow([
            'Entrepôt',
            'Code Article',
            'Nom Article',
            'Unité',
            'Quantité',
            'Coût Moyen (MAD)',
            'Valeur Totale (MAD)'
        ])

        # Write product data by warehouse
        for warehouse in valuation_data['warehouses']:
            for product in warehouse['products']:
                writer.writerow([
                    warehouse['warehouse_name'],
                    product['product_sku'],
                    product['product_name'],
                    product['uom'],
                    f"{product['quantity']:.2f}",
                    f"{product['average_cost']:.2f}",
                    f"{product['total_value']:.2f}"
                ])

        # Write summary
        writer.writerow([])
        writer.writerow([
            'TOTAL GÉNÉRAL',
            '',
            '',
            '',
            f"{valuation_data['summary']['total_quantity']:.2f}",
            '',
            f"{valuation_data['summary']['total_value']:.2f}"
        ])

        return output.getvalue()

    def export_vat_sales_to_csv(self, vat_data: Dict[str, Any]) -> str:
        """Export Sales by VAT Rate report to CSV with French headers."""
        import csv
        import io

        output = io.StringIO()
        writer = csv.writer(output)

        # French headers for VAT Sales
        writer.writerow([
            'Taux TVA (%)',
            'Base HT (MAD)',
            'Montant TVA (MAD)',
            'Total TTC (MAD)',
            'Nombre de Transactions'
        ])

        # Write VAT rate data
        for rate_info in vat_data['vat_rates']:
            writer.writerow([
                f"{rate_info['vat_rate']:.2f}%",
                f"{rate_info['base_amount']:.2f}",
                f"{rate_info['vat_amount']:.2f}",
                f"{rate_info['total_amount']:.2f}",
                rate_info['transaction_count']
            ])

        # Write summary
        writer.writerow([])
        writer.writerow([
            'TOTAL GÉNÉRAL',
            f"{vat_data['summary']['total_base_amount']:.2f}",
            f"{vat_data['summary']['total_vat_amount']:.2f}",
            f"{vat_data['summary']['total_amount']:.2f}",
            ''
        ])

        return output.getvalue()

    def export_to_xlsx(self, data: Dict[str, Any], report_type: str) -> bytes:
        """Export report data to XLSX format with French headers."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl is required for XLSX export")

        wb = openpyxl.Workbook()
        ws = wb.active

        # Set worksheet title based on report type
        report_titles = {
            'aged_ar': 'Comptes Clients Âgés',
            'aged_ap': 'Comptes Fournisseurs Âgés',
            'stock_valuation': 'Valorisation Stock',
            'vat_sales': 'Ventes par Taux TVA',
            'vat_purchases': 'Achats par Taux TVA'
        }

        ws.title = report_titles.get(report_type, 'Rapport')

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        if report_type == 'aged_ar':
            headers = ['Client', 'Courant (MAD)', '1-30 jours (MAD)', '31-60 jours (MAD)',
                      '61-90 jours (MAD)', 'Plus de 90 jours (MAD)', 'Total (MAD)']
            ws.append(headers)

            # Style headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Add data
            for customer in data['customers']:
                ws.append([
                    customer['customer_name'],
                    float(customer['current']),
                    float(customer['1_30_days']),
                    float(customer['31_60_days']),
                    float(customer['61_90_days']),
                    float(customer['over_90_days']),
                    float(customer['total'])
                ])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        import io
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()
